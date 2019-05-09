# -*- coding: UTF-8 -*-
'''
Created on 2019年4月27日

@author: 13892
'''

from openpyxl import load_workbook
wb = load_workbook(filename = 'sample.xlsx')
sheet_ranges = wb['Sheet1']
for i in sheet_ranges.iter_cols(min_col=50,max_col=50):
    print(i)
print(type(sheet_ranges))
print(sheet_ranges['1'])       
print(sheet_ranges['A'])       
print(sheet_ranges['A1'].font)       
print(sheet_ranges['B1'].font)